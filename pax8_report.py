#!/usr/bin/env python3
"""
PAX8 Microsoft License Report Generator
Generates an Excel report of all Microsoft subscriptions across all clients,
including full subscription history.
"""

import os
import sys
import time
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

API_BASE = "https://api.pax8.com/v1"
TOKEN_URL = "https://token-manager.pax8.com/oauth/token"
PAGE_SIZE = 200  # max allowed by PAX8 API
MAX_RETRIES = 5
INITIAL_BACKOFF = 1  # seconds


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_credentials():
    """Load and validate PAX8 credentials from .env file."""
    load_dotenv()
    client_id = os.getenv("PAX8_CLIENT_ID", "").strip()
    client_secret = os.getenv("PAX8_CLIENT_SECRET", "").strip()

    if not client_id or not client_secret or client_id == "your_client_id_here":
        print("\nERROR: PAX8 credentials not configured.")
        print("  1. Copy .env.example to .env")
        print("  2. Replace the placeholder values with your real PAX8 API credentials.")
        print("  3. Credentials can be found at: https://app.pax8.com/integrations/credentials")
        sys.exit(1)

    return client_id, client_secret


def api_request(session, method, url, retries=MAX_RETRIES, **kwargs):
    """Make an API request with retry logic and exponential backoff."""
    backoff = INITIAL_BACKOFF
    last_error = None

    for attempt in range(retries):
        try:
            resp = session.request(method, url, **kwargs)

            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", backoff))
                print(f"    Rate limited. Retrying in {retry_after}s...")
                time.sleep(retry_after)
                backoff *= 2
                continue

            resp.raise_for_status()
            return resp.json()

        except requests.exceptions.HTTPError as e:
            last_error = e
            if resp.status_code >= 500:
                print(f"    Server error ({resp.status_code}). Retry {attempt + 1}/{retries} in {backoff}s...")
                time.sleep(backoff)
                backoff *= 2
                continue
            raise
        except requests.exceptions.ConnectionError as e:
            last_error = e
            print(f"    Connection error. Retry {attempt + 1}/{retries} in {backoff}s...")
            time.sleep(backoff)
            backoff *= 2
            continue

    raise last_error


def paginate(session, url, params=None):
    """Fetch all pages from a paginated PAX8 endpoint."""
    if params is None:
        params = {}
    params["size"] = PAGE_SIZE
    params["page"] = 0

    all_items = []
    while True:
        data = api_request(session, "GET", url, params=params)
        content = data.get("content", [])
        all_items.extend(content)

        page_info = data.get("page", {})
        current_page = page_info.get("number", 0)
        total_pages = page_info.get("totalPages", 1)

        if current_page + 1 >= total_pages:
            break
        params["page"] = current_page + 1

    return all_items


# ---------------------------------------------------------------------------
# API Calls
# ---------------------------------------------------------------------------

def authenticate(client_id, client_secret):
    """Authenticate with PAX8 and return a configured requests session."""
    print("Authenticating with PAX8 API...", end=" ", flush=True)

    try:
        resp = requests.post(TOKEN_URL, json={
            "client_id": client_id,
            "client_secret": client_secret,
            "audience": "https://api.pax8.com",
            "grant_type": "client_credentials",
        })
        resp.raise_for_status()
    except requests.exceptions.HTTPError:
        print("FAILED.")
        print("\nERROR: Authentication failed. Please check your .env credentials.")
        print("  Ensure PAX8_CLIENT_ID and PAX8_CLIENT_SECRET are correct.")
        print(f"  HTTP {resp.status_code}: {resp.text}")
        sys.exit(1)
    except requests.exceptions.ConnectionError:
        print("FAILED.")
        print("\nERROR: Could not connect to PAX8 authentication server.")
        print("  Please check your internet connection.")
        sys.exit(1)

    token = resp.json().get("access_token")
    if not token:
        print("FAILED.")
        print("\nERROR: No access token in authentication response.")
        sys.exit(1)

    session = requests.Session()
    session.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    })
    print("Success.")
    return session


def fetch_companies(session):
    """Fetch all companies (clients) from PAX8."""
    print("Fetching companies...", end=" ", flush=True)
    companies = paginate(session, f"{API_BASE}/companies")
    print(f"Found {len(companies)} clients.")
    return companies


def fetch_subscriptions_for_company(session, company_id, statuses=None):
    """Fetch subscriptions for a company, optionally filtering by status list."""
    if statuses is None:
        statuses = ["Active", "Cancelled", "PendingCancel"]

    all_subs = []
    for status in statuses:
        subs = paginate(session, f"{API_BASE}/subscriptions", params={
            "companyId": company_id,
            "status": status,
        })
        all_subs.extend(subs)
    return all_subs


def fetch_product(session, product_id, product_cache):
    """Fetch product details (with caching)."""
    if product_id in product_cache:
        return product_cache[product_id]

    try:
        product = api_request(session, "GET", f"{API_BASE}/products/{product_id}")
        product_cache[product_id] = product
        return product
    except Exception:
        product_cache[product_id] = None
        return None


def fetch_subscription_history(session, subscription_id):
    """Fetch history for a subscription."""
    try:
        data = api_request(session, "GET", f"{API_BASE}/subscriptions/{subscription_id}/history")
        return data.get("content", data) if isinstance(data, dict) else data
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Excel Report Generation
# ---------------------------------------------------------------------------

def style_header_row(ws, num_cols):
    """Apply dark header style to the first row."""
    header_fill = PatternFill(start_color="2F3640", end_color="2F3640", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="555555"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border


def auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 50)


def apply_alternating_rows(ws, num_cols):
    """Apply alternating row shading."""
    light_fill = PatternFill(start_color="F5F6FA", end_color="F5F6FA", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=col).fill = light_fill


def format_date(value):
    """Format a date string to YYYY-MM-DD."""
    if not value:
        return ""
    try:
        if "T" in str(value):
            return str(value).split("T")[0]
        return str(value)[:10]
    except Exception:
        return str(value)


def generate_report(summary_rows, history_rows):
    """Generate the Excel report and return the file path."""
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"pax8_microsoft_license_report_{today}.xlsx"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

    wb = Workbook()

    # --- Tab 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = [
        "Company Name", "Subscription ID", "Product ID", "Product Name", "SKU",
        "Status", "Current Quantity", "Start Date", "Billing Term",
        "Commitment Term", "Price Per Unit",
    ]
    ws_summary.append(summary_headers)

    # Sort by company name
    summary_rows.sort(key=lambda r: r[0].lower())
    for row in summary_rows:
        ws_summary.append(row)

    style_header_row(ws_summary, len(summary_headers))
    apply_alternating_rows(ws_summary, len(summary_headers))
    auto_fit_columns(ws_summary)
    ws_summary.auto_filter.ref = ws_summary.dimensions

    # --- Tab 2: Subscription History ---
    ws_history = wb.create_sheet("Subscription History")
    history_headers = [
        "Company Name", "Subscription ID", "Product Name", "Date",
        "Action/Change Type", "Quantity Change", "Total Quantity After Change",
    ]
    ws_history.append(history_headers)

    # Sort by company name, then date
    history_rows.sort(key=lambda r: (r[0].lower(), r[3]))
    for row in history_rows:
        ws_history.append(row)

    style_header_row(ws_history, len(history_headers))
    apply_alternating_rows(ws_history, len(history_headers))
    auto_fit_columns(ws_history)
    ws_history.auto_filter.ref = ws_history.dimensions

    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    client_id, client_secret = load_credentials()
    session = authenticate(client_id, client_secret)

    companies = fetch_companies(session)
    if not companies:
        print("No companies found. Exiting.")
        sys.exit(0)

    product_cache = {}
    summary_rows = []
    history_rows = []
    errors = []
    total_subscriptions = 0
    total_history_records = 0

    for idx, company in enumerate(companies, 1):
        company_name = company.get("name", "Unknown")
        company_id = company.get("id")

        try:
            subs = fetch_subscriptions_for_company(session, company_id)
        except Exception as e:
            error_msg = f"Failed to fetch subscriptions for {company_name}: {e}"
            print(f"  ERROR: {error_msg}")
            errors.append(error_msg)
            continue

        # Filter to Microsoft subscriptions by looking up each product
        ms_subs = []
        for sub in subs:
            product_id = sub.get("productId")
            if not product_id:
                continue
            product = fetch_product(session, product_id, product_cache)
            if product and "microsoft" in product.get("vendorName", "").lower():
                ms_subs.append((sub, product))

        print(f"Processing client {idx}/{len(companies)}: {company_name} — {len(ms_subs)} Microsoft subscriptions found")

        total_subscriptions += len(ms_subs)

        # Diagnostic: dump the first subscription's billing fields for debugging
        if ms_subs and idx == 1:
            sample_sub = ms_subs[0][0]
            print("\n  --- DIAGNOSTIC: Raw subscription fields (first sub, first client) ---")
            for field in ("billingTerm", "commitmentTerm", "price", "partnerCost",
                          "billingStart", "startDate", "endDate", "status", "quantity"):
                print(f"    {field}: {sample_sub.get(field)!r}")
            print("  --- END DIAGNOSTIC ---\n")

        for sub_idx, (sub, product) in enumerate(ms_subs, 1):
            sub_id = sub.get("id", "")
            product_id = sub.get("productId", "")
            product_name = product.get("name", "")
            sku = product.get("sku", "") or product.get("vendorSku", "")
            status = sub.get("status", "")
            quantity = sub.get("quantity", 0)
            start_date = format_date(sub.get("startDate"))
            billing_term = sub.get("billingTerm", "")
            commitment = sub.get("commitmentTerm")
            if isinstance(commitment, dict) and commitment:
                commitment_term = commitment.get("term", "")
                commitment_end = format_date(commitment.get("endDate"))
                commitment_str = f"{commitment_term} (ends {commitment_end})" if commitment_end else commitment_term
            else:
                commitment_str = ""
            price = sub.get("price", "")

            summary_rows.append([
                company_name, sub_id, product_id, product_name, sku,
                status, quantity, start_date, billing_term,
                commitment_str, price,
            ])

            # Fetch subscription history
            print(f"  Fetching history for subscription {sub_idx}/{len(ms_subs)}...", flush=True)
            try:
                history = fetch_subscription_history(session, sub_id)
            except Exception as e:
                error_msg = f"Failed to fetch history for subscription {sub_id} ({company_name}): {e}"
                print(f"    ERROR: {error_msg}")
                errors.append(error_msg)
                history = []

            if isinstance(history, list):
                prev_quantity = None
                # Sort history by date ascending for change calculation
                sorted_history = sorted(history, key=lambda h: h.get("createdDate", "") or "")

                for record in sorted_history:
                    record_date = format_date(record.get("createdDate"))
                    record_qty = record.get("quantity", 0)
                    record_status = record.get("status", "")

                    if prev_quantity is not None:
                        qty_change = record_qty - prev_quantity
                        change_str = f"+{qty_change}" if qty_change > 0 else str(qty_change)
                    else:
                        change_str = "Initial"
                        qty_change = record_qty

                    # Determine action type
                    if prev_quantity is None:
                        action = "Initial Provisioning"
                    elif record_status == "Cancelled":
                        action = "Cancellation"
                    elif qty_change > 0:
                        action = "Quantity Increase"
                    elif qty_change < 0:
                        action = "Quantity Decrease"
                    else:
                        action = "Status/Config Change"

                    history_rows.append([
                        company_name, sub_id, product_name, record_date,
                        action, change_str, record_qty,
                    ])
                    total_history_records += 1
                    prev_quantity = record_qty

    # Generate the Excel report
    print("\nGenerating Excel report...", end=" ", flush=True)
    filepath = generate_report(summary_rows, history_rows)
    print("Done.")

    # Final summary
    print("\n" + "=" * 60)
    print("REPORT COMPLETE")
    print("=" * 60)
    print(f"  Clients processed:       {len(companies)}")
    print(f"  Microsoft subscriptions: {total_subscriptions}")
    print(f"  History records:         {total_history_records}")
    if errors:
        print(f"  Errors encountered:      {len(errors)}")
        print("\n  Errors:")
        for err in errors:
            print(f"    - {err}")
    else:
        print(f"  Errors encountered:      0")
    print(f"\nReport saved to: {os.path.abspath(filepath)}")


if __name__ == "__main__":
    main()
