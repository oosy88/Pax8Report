#!/usr/bin/env python3
"""
PAX8 Microsoft License Optimization Analyzer
Analyzes Microsoft license trends from the PAX8 report and recommends
annual vs monthly commitment splits to reduce costs.

Fetches live pricing from the PAX8 API for each product found in the report.
"""

import glob
import math
import os
import sys
import time
from collections import defaultdict
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

API_BASE = "https://api.pax8.com/v1"
TOKEN_URL = "https://token-manager.pax8.com/oauth/token"
MAX_RETRIES = 5
INITIAL_BACKOFF = 1

SUMMARY_REQUIRED_COLS = [
    "Company Name", "Subscription ID", "Product ID", "Product Name", "SKU",
    "Status", "Current Quantity", "Start Date", "Billing Term",
]
HISTORY_REQUIRED_COLS = [
    "Company Name", "Subscription ID", "Product Name", "Date",
    "Action/Change Type", "Quantity Change", "Total Quantity After Change",
]


# ---------------------------------------------------------------------------
# PAX8 API — Authentication & Pricing
# ---------------------------------------------------------------------------

def load_credentials():
    """Load PAX8 credentials from .env."""
    load_dotenv()
    client_id = os.getenv("PAX8_CLIENT_ID", "").strip()
    client_secret = os.getenv("PAX8_CLIENT_SECRET", "").strip()

    if not client_id or not client_secret or client_id == "your_client_id_here":
        print("\nWARNING: PAX8 credentials not configured in .env")
        print("  Dollar savings calculations will be skipped.")
        print("  To enable pricing, add PAX8_CLIENT_ID and PAX8_CLIENT_SECRET to .env")
        return None, None

    return client_id, client_secret


def authenticate(client_id, client_secret):
    """Authenticate with PAX8 and return a configured requests session."""
    print("Authenticating with PAX8 API...", end=" ", flush=True)

    try:
        resp = requests.post(TOKEN_URL, json={
            "client_id": client_id,
            "client_secret": client_secret,
            "audience": "https://api.pax8.com",
            "grant_type": "client_credentials",
        })
        resp.raise_for_status()
    except requests.exceptions.HTTPError:
        print("FAILED.")
        print(f"  Authentication failed (HTTP {resp.status_code}). Check .env credentials.")
        return None
    except requests.exceptions.ConnectionError:
        print("FAILED.")
        print("  Could not connect to PAX8. Check your internet connection.")
        return None

    token = resp.json().get("access_token")
    if not token:
        print("FAILED.")
        print("  No access token in response.")
        return None

    session = requests.Session()
    session.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    })
    print("Success.")
    return session


def api_request(session, method, url, retries=MAX_RETRIES, **kwargs):
    """Make an API request with retry logic and exponential backoff."""
    backoff = INITIAL_BACKOFF
    last_error = None

    for attempt in range(retries):
        try:
            resp = session.request(method, url, **kwargs)

            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", backoff))
                print(f"    Rate limited. Retrying in {retry_after}s...")
                time.sleep(retry_after)
                backoff *= 2
                continue

            resp.raise_for_status()
            return resp.json()

        except requests.exceptions.HTTPError as e:
            last_error = e
            if resp.status_code >= 500:
                print(f"    Server error ({resp.status_code}). Retry {attempt + 1}/{retries} in {backoff}s...")
                time.sleep(backoff)
                backoff *= 2
                continue
            raise
        except requests.exceptions.ConnectionError as e:
            last_error = e
            print(f"    Connection error. Retry {attempt + 1}/{retries} in {backoff}s...")
            time.sleep(backoff)
            backoff *= 2
            continue

    raise last_error


def fetch_product_pricing(session, product_id):
    """
    Fetch pricing for a product from the PAX8 API.
    Returns dict with monthly_price and annual_price, or None if unavailable.
    """
    try:
        data = api_request(session, "GET", f"{API_BASE}/products/{product_id}/pricing")
    except Exception:
        return None

    content = data.get("content", []) if isinstance(data, dict) else data
    if not content:
        return None

    monthly_rate = None
    annual_rate = None

    for entry in content:
        billing_term = (entry.get("billingTerm") or "").strip()
        rates = entry.get("rates", [])
        if not rates:
            continue

        # Use the first rate's partnerBuyRate (our cost)
        rate = rates[0].get("partnerBuyRate")
        if rate is None:
            continue

        if billing_term == "Monthly":
            monthly_rate = float(rate)
        elif billing_term == "Annual":
            annual_rate = float(rate)

    if monthly_rate is not None and annual_rate is not None:
        return {"monthly_price": monthly_rate, "annual_price": annual_rate}

    # If only one term exists, we can't calculate savings between the two
    return None


def fetch_all_pricing(session, product_ids):
    """
    Fetch pricing for all unique product IDs.
    Returns dict: product_id -> {monthly_price, annual_price}
    """
    pricing = {}
    total = len(product_ids)
    fetched = 0
    failed = []

    for i, pid in enumerate(product_ids, 1):
        print(f"  Fetching pricing {i}/{total}: {pid[:12]}...", end=" ", flush=True)
        result = fetch_product_pricing(session, pid)
        if result:
            pricing[pid] = result
            fetched += 1
            print(f"${result['monthly_price']:.2f}/mo, ${result['annual_price']:.2f}/yr")
        else:
            failed.append(pid)
            print("no monthly+annual pricing available")

    print(f"  Pricing retrieved for {fetched}/{total} products")
    if failed:
        print(f"  {len(failed)} products missing monthly+annual pricing comparison")

    return pricing


# ---------------------------------------------------------------------------
# Data Ingestion
# ---------------------------------------------------------------------------

def find_input_file(args):
    """Determine the input Excel file from args or auto-detect."""
    if len(args) > 1:
        path = args[1]
        if not os.path.isfile(path):
            print(f"ERROR: File not found: {path}")
            sys.exit(1)
        return path

    pattern = "pax8_microsoft_license_report_*.xlsx"
    matches = sorted(glob.glob(pattern), reverse=True)
    if not matches:
        print("ERROR: No input file specified and no pax8_microsoft_license_report_*.xlsx found in current directory.")
        print("Usage: python3 license_optimizer.py <report_file.xlsx>")
        sys.exit(1)

    print(f"Auto-detected input file: {matches[0]}")
    return matches[0]


def read_sheet(wb, sheet_name, required_cols):
    """Read a sheet into a list of dicts, validating required columns."""
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found in workbook.")
        print(f"  Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"ERROR: Sheet '{sheet_name}' is empty.")
        sys.exit(1)

    headers = [str(h).strip() if h else "" for h in rows[0]]
    missing = [c for c in required_cols if c not in headers]
    if missing:
        print(f"ERROR: Sheet '{sheet_name}' is missing required columns: {missing}")
        print(f"  Found columns: {headers}")
        sys.exit(1)

    records = []
    for row in rows[1:]:
        record = {}
        for i, header in enumerate(headers):
            record[header] = row[i] if i < len(row) else None
        records.append(record)
    return records


# ---------------------------------------------------------------------------
# Timeline Reconstruction
# ---------------------------------------------------------------------------

def month_key(dt):
    """Return 'YYYY-MM' string from a date/datetime."""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m")
    if isinstance(dt, str) and len(dt) >= 7:
        return dt[:7]
    return None


def months_between(start_ym, end_ym):
    """Generate list of 'YYYY-MM' strings from start to end inclusive."""
    sy, sm = int(start_ym[:4]), int(start_ym[5:7])
    ey, em = int(end_ym[:4]), int(end_ym[5:7])
    result = []
    y, m = sy, sm
    while (y, m) <= (ey, em):
        result.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return result


def build_timelines(summary_records, history_records):
    """
    Build month-by-month license count timelines.
    Returns dict: (company, product) -> {sku, product_id, timeline, current_count, status}
    """
    combo_info = {}
    for rec in summary_records:
        company = str(rec.get("Company Name") or "").strip()
        product = str(rec.get("Product Name") or "").strip()
        if not company or not product:
            continue
        key = (company, product)
        sku = str(rec.get("SKU") or "").strip()
        product_id = str(rec.get("Product ID") or "").strip()
        qty = _to_int(rec.get("Current Quantity", 0))
        status = str(rec.get("Status") or "").strip()

        if key not in combo_info:
            combo_info[key] = {
                "sku": sku,
                "product_id": product_id,
                "current_count": 0,
                "status": status,
            }

        combo_info[key]["current_count"] += qty
        # Keep the product_id if we have one
        if product_id and not combo_info[key]["product_id"]:
            combo_info[key]["product_id"] = product_id

    # Per-subscription history tracking
    sub_history = defaultdict(list)
    for rec in history_records:
        company = str(rec.get("Company Name") or "").strip()
        product = str(rec.get("Product Name") or "").strip()
        sub_id = str(rec.get("Subscription ID") or "").strip()
        date_str = str(rec.get("Date") or "").strip()
        qty_after = _to_int(rec.get("Total Quantity After Change", 0))
        if not company or not product or not date_str:
            continue
        mk = month_key(date_str)
        if mk:
            sub_history[(company, product, sub_id)].append((mk, qty_after))

    today_ym = datetime.now().strftime("%Y-%m")

    for key, info in combo_info.items():
        company, product = key

        related_subs = {k: v for k, v in sub_history.items()
                        if k[0] == company and k[1] == product}

        if not related_subs:
            info["timeline"] = {today_ym: info["current_count"]}
            continue

        all_months = set()
        for events in related_subs.values():
            for mk, _ in events:
                all_months.add(mk)
        if not all_months:
            info["timeline"] = {today_ym: info["current_count"]}
            continue

        min_month = min(all_months)
        max_month = max(max(all_months), today_ym)
        month_range = months_between(min_month, max_month)

        sub_timelines = {}
        for sub_key, events in related_subs.items():
            sub_id = sub_key[2]
            monthly = {}
            for mk, qty in sorted(events):
                monthly[mk] = qty
            tl = {}
            last_val = 0
            for m in month_range:
                if m in monthly:
                    last_val = monthly[m]
                tl[m] = last_val
            sub_timelines[sub_id] = tl

        timeline = {}
        for m in month_range:
            total = sum(tl.get(m, 0) for tl in sub_timelines.values())
            timeline[m] = total

        info["timeline"] = timeline

    return combo_info


def _to_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


# ---------------------------------------------------------------------------
# Trend Metrics
# ---------------------------------------------------------------------------

def calculate_metrics(timeline, current_count):
    """Calculate trend metrics from a month-by-month timeline dict."""
    if not timeline:
        return empty_metrics(current_count)

    sorted_months = sorted(timeline.keys())
    values = [timeline[m] for m in sorted_months]
    n = len(values)

    if n == 0:
        return empty_metrics(current_count)

    trail_12 = values[-12:] if n >= 12 else values
    trail_6 = values[-6:] if n >= 6 else values

    min_12 = min(trail_12)
    min_6 = min(trail_6)
    max_12 = max(trail_12)

    if len(trail_12) > 1:
        mean = sum(trail_12) / len(trail_12)
        variance = sum((x - mean) ** 2 for x in trail_12) / (len(trail_12) - 1)
        std_dev = math.sqrt(variance)
    else:
        std_dev = 0.0

    slope = _linear_slope(trail_12)

    if current_count > 0:
        slope_pct = slope / current_count
    else:
        slope_pct = 0
    if slope_pct > 0.005:
        trend = "Growing"
    elif slope_pct < -0.005:
        trend = "Shrinking"
    else:
        trend = "Stable"

    months_since_decrease = None
    for i in range(len(values) - 1, 0, -1):
        if values[i] < values[i - 1]:
            months_since_decrease = len(values) - 1 - i
            break
    if months_since_decrease is None:
        months_since_decrease = len(values)

    largest_drop = 0
    for i in range(1, len(values)):
        drop = values[i - 1] - values[i]
        if drop > largest_drop:
            largest_drop = drop

    return {
        "current_count": current_count,
        "min_12": min_12,
        "min_6": min_6,
        "max_12": max_12,
        "std_dev": round(std_dev, 2),
        "slope": round(slope, 4),
        "trend": trend,
        "months_since_decrease": months_since_decrease,
        "largest_drop": largest_drop,
        "months_of_data": n,
    }


def empty_metrics(current_count):
    return {
        "current_count": current_count,
        "min_12": current_count,
        "min_6": current_count,
        "max_12": current_count,
        "std_dev": 0.0,
        "slope": 0.0,
        "trend": "Stable",
        "months_since_decrease": 0,
        "largest_drop": 0,
        "months_of_data": 0,
    }


def _linear_slope(values):
    n = len(values)
    if n < 2:
        return 0.0
    x_mean = (n - 1) / 2.0
    y_mean = sum(values) / n
    numerator = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(values))
    denominator = sum((i - x_mean) ** 2 for i in range(n))
    if denominator == 0:
        return 0.0
    return numerator / denominator


# ---------------------------------------------------------------------------
# Recommendations
# ---------------------------------------------------------------------------

def generate_recommendations(metrics):
    cc = metrics["current_count"]
    min12 = metrics["min_12"]
    min6 = metrics["min_6"]
    trend = metrics["trend"]
    months_since_dec = metrics["months_since_decrease"]
    std_dev = metrics["std_dev"]
    months_data = metrics["months_of_data"]

    notes = []

    if cc <= 1:
        notes.append("Single license — keep monthly")
        return {
            "conservative": 0, "moderate": 0, "aggressive": 0,
            "notes": "; ".join(notes),
        }

    if months_data < 3:
        notes.append("Low confidence — limited history")

    conservative = max(0, math.floor(min12 * 0.9))

    if trend in ("Stable", "Growing"):
        moderate = min12
    else:
        moderate = conservative

    low_volatility = (std_dev < 0.1 * cc) if cc > 0 else True
    if trend == "Growing" and months_since_dec >= 6 and low_volatility:
        aggressive = min6
    else:
        aggressive = moderate

    conservative = min(conservative, cc)
    moderate = min(moderate, cc)
    aggressive = min(aggressive, cc)

    return {
        "conservative": conservative,
        "moderate": moderate,
        "aggressive": aggressive,
        "notes": "; ".join(notes),
    }


def calculate_savings(rec, metrics, price_info):
    cc = metrics["current_count"]
    min12 = metrics["min_12"]

    if price_info is None:
        return {tier: {"savings": None, "risk": None, "total_cost": None, "current_cost": None}
                for tier in ("conservative", "moderate", "aggressive")}

    mp = price_info["monthly_price"]
    ap = price_info["annual_price"]
    current_annual_cost = mp * cc * 12

    result = {}
    for tier in ("conservative", "moderate", "aggressive"):
        annual_qty = rec[tier]
        monthly_qty = cc - annual_qty

        annual_cost = ap * annual_qty * 12
        monthly_cost = mp * monthly_qty * 12
        total_cost = annual_cost + monthly_cost
        net_savings = current_annual_cost - total_cost

        overpay_qty = max(0, annual_qty - min12)
        risk = overpay_qty * ap * 6

        result[tier] = {
            "savings": round(net_savings, 2),
            "risk": round(risk, 2),
            "total_cost": round(total_cost, 2),
            "current_cost": round(current_annual_cost, 2),
        }
    return result


# ---------------------------------------------------------------------------
# Excel Generation
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="2F3640", end_color="2F3640", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="F5F6FA", end_color="F5F6FA", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TOTALS_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
TOTALS_FONT = Font(bold=True, size=11)
CURRENCY_FMT = '#,##0.00'


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def auto_fit(ws):
    for col_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 3, 45)


def alt_rows(ws, num_cols):
    for r in range(2, ws.max_row + 1):
        if r % 2 == 0:
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).fill = ALT_FILL


def freeze_panes(ws, row=2, col=3):
    ws.freeze_panes = ws.cell(row=row, column=col)


def fmt_currency(ws, col_indices):
    for r in range(2, ws.max_row + 1):
        for c in col_indices:
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = CURRENCY_FMT


def write_recommendations_tab(wb, results, has_pricing):
    ws = wb.active
    ws.title = "Recommendations"

    headers = [
        "Company Name", "Product Name", "SKU", "Current License Count",
        "12-Month Min", "6-Month Min", "Trend Direction",
        "Months Since Last Decrease", "Volatility (Std Dev)",
        "Conservative Annual Qty", "Conservative Monthly Qty",
    ]
    currency_cols = []
    if has_pricing:
        headers += ["Conservative Annual Savings", "Conservative Risk Exposure"]
        currency_cols += [12, 13]
    headers += ["Moderate Annual Qty", "Moderate Monthly Qty"]
    if has_pricing:
        headers += ["Moderate Annual Savings", "Moderate Risk Exposure"]
        currency_cols += [len(headers) - 1, len(headers)]
    headers += ["Aggressive Annual Qty", "Aggressive Monthly Qty"]
    if has_pricing:
        headers += ["Aggressive Annual Savings", "Aggressive Risk Exposure"]
        currency_cols += [len(headers) - 1, len(headers)]
    headers.append("Notes")

    ws.append(headers)

    sorted_keys = sorted(results.keys(), key=lambda k: (k[0].lower(), k[1].lower()))
    trend_col = headers.index("Trend Direction") + 1

    for key in sorted_keys:
        r = results[key]
        m = r["metrics"]
        rec = r["rec"]
        sav = r["savings"]
        cc = m["current_count"]

        row = [
            key[0], key[1], r["sku"], cc,
            m["min_12"], m["min_6"], m["trend"],
            m["months_since_decrease"], m["std_dev"],
            rec["conservative"], cc - rec["conservative"],
        ]
        if has_pricing:
            cs = sav["conservative"]
            row += [cs["savings"], cs["risk"]]
        row += [rec["moderate"], cc - rec["moderate"]]
        if has_pricing:
            ms = sav["moderate"]
            row += [ms["savings"], ms["risk"]]
        row += [rec["aggressive"], cc - rec["aggressive"]]
        if has_pricing:
            ags = sav["aggressive"]
            row += [ags["savings"], ags["risk"]]
        row.append(rec["notes"])

        ws.append(row)

    num_cols = len(headers)
    style_header(ws, num_cols)
    alt_rows(ws, num_cols)

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=trend_col)
        if cell.value == "Growing":
            cell.fill = GREEN_FILL
        elif cell.value == "Stable":
            cell.fill = YELLOW_FILL
        elif cell.value == "Shrinking":
            cell.fill = RED_FILL

    if currency_cols:
        fmt_currency(ws, currency_cols)

    auto_fit(ws)
    ws.auto_filter.ref = ws.dimensions
    freeze_panes(ws)


def write_trends_tab(wb, results, combo_info):
    ws = wb.create_sheet("Client Trends")

    all_months = set()
    for info in combo_info.values():
        all_months.update(info.get("timeline", {}).keys())

    if not all_months:
        ws.append(["No timeline data available."])
        return

    sorted_months = sorted(all_months)
    headers = ["Company Name", "Product Name"] + sorted_months
    ws.append(headers)

    sorted_keys = sorted(combo_info.keys(), key=lambda k: (k[0].lower(), k[1].lower()))
    for key in sorted_keys:
        info = combo_info[key]
        tl = info.get("timeline", {})
        row = [key[0], key[1]] + [tl.get(m, "") for m in sorted_months]
        ws.append(row)

    num_cols = len(headers)
    style_header(ws, num_cols)
    alt_rows(ws, num_cols)
    auto_fit(ws)
    ws.auto_filter.ref = ws.dimensions
    freeze_panes(ws)


def write_savings_summary_tab(wb, results, has_pricing):
    ws = wb.create_sheet("Savings Summary")

    headers = [
        "Company Name", "Total Current Annual Spend",
        "Conservative Total Savings", "Moderate Total Savings",
        "Aggressive Total Savings", "Number of Products Analyzed",
        "Overall Trend Assessment",
    ]
    ws.append(headers)

    company_data = defaultdict(lambda: {
        "current_cost": 0, "con_sav": 0, "mod_sav": 0, "agg_sav": 0,
        "num_products": 0, "trends": [],
    })

    for key, r in results.items():
        company = key[0]
        d = company_data[company]
        d["num_products"] += 1
        d["trends"].append(r["metrics"]["trend"])
        sav = r["savings"]
        if has_pricing and sav["conservative"]["current_cost"] is not None:
            d["current_cost"] += sav["conservative"]["current_cost"]
            d["con_sav"] += sav["conservative"]["savings"] or 0
            d["mod_sav"] += sav["moderate"]["savings"] or 0
            d["agg_sav"] += sav["aggressive"]["savings"] or 0

    sorted_companies = sorted(company_data.keys(), key=str.lower)
    totals = {"current_cost": 0, "con_sav": 0, "mod_sav": 0, "agg_sav": 0, "num_products": 0}

    for company in sorted_companies:
        d = company_data[company]
        trend_counts = defaultdict(int)
        for t in d["trends"]:
            trend_counts[t] += 1
        dominant = max(trend_counts, key=trend_counts.get)
        if len(trend_counts) > 1 and trend_counts[dominant] < len(d["trends"]):
            overall = "Mixed"
        else:
            overall = dominant

        row = [
            company,
            d["current_cost"] if has_pricing else "N/A",
            d["con_sav"] if has_pricing else "N/A",
            d["mod_sav"] if has_pricing else "N/A",
            d["agg_sav"] if has_pricing else "N/A",
            d["num_products"],
            overall,
        ]
        ws.append(row)

        totals["current_cost"] += d["current_cost"]
        totals["con_sav"] += d["con_sav"]
        totals["mod_sav"] += d["mod_sav"]
        totals["agg_sav"] += d["agg_sav"]
        totals["num_products"] += d["num_products"]

    totals_row = [
        "TOTAL",
        totals["current_cost"] if has_pricing else "N/A",
        totals["con_sav"] if has_pricing else "N/A",
        totals["mod_sav"] if has_pricing else "N/A",
        totals["agg_sav"] if has_pricing else "N/A",
        totals["num_products"],
        "",
    ]
    ws.append(totals_row)

    totals_row_idx = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=totals_row_idx, column=c)
        cell.fill = TOTALS_FILL
        cell.font = TOTALS_FONT

    num_cols = len(headers)
    style_header(ws, num_cols)
    alt_rows(ws, num_cols)
    if has_pricing:
        fmt_currency(ws, [2, 3, 4, 5])
    auto_fit(ws)
    ws.auto_filter.ref = ws.dimensions


def write_unpriced_tab(wb, unpriced):
    """Tab 4: Products where PAX8 API didn't return both monthly+annual pricing."""
    ws = wb.create_sheet("Unpriced Products")
    headers = ["Product Name", "SKU", "Product ID", "Number of Clients Using It",
               "Total Licenses Across All Clients", "Reason"]
    ws.append(headers)

    sorted_products = sorted(unpriced.items(), key=lambda x: x[1]["total_licenses"], reverse=True)
    for (name, sku, pid), info in sorted_products:
        ws.append([name, sku, pid, info["client_count"], info["total_licenses"], info["reason"]])

    num_cols = len(headers)
    style_header(ws, num_cols)
    alt_rows(ws, num_cols)
    auto_fit(ws)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    try:
        _run()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(1)
    except Exception as e:
        print(f"\nERROR: {e}")
        sys.exit(1)


def _run():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = find_input_file(sys.argv)

    print(f"Reading input file: {input_file}")
    try:
        wb = load_workbook(input_file, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR: Could not open Excel file: {e}")
        sys.exit(1)

    summary_records = read_sheet(wb, "Summary", SUMMARY_REQUIRED_COLS)
    history_records = read_sheet(wb, "Subscription History", HISTORY_REQUIRED_COLS)
    wb.close()

    # Count unique combos and collect product IDs
    combos = set()
    product_id_map = {}  # product_id -> (product_name, sku)
    for rec in summary_records:
        company = str(rec.get("Company Name") or "").strip()
        product = str(rec.get("Product Name") or "").strip()
        product_id = str(rec.get("Product ID") or "").strip()
        sku = str(rec.get("SKU") or "").strip()
        if company and product:
            combos.add((company, product))
        if product_id and product_id not in product_id_map:
            product_id_map[product_id] = (product, sku)

    companies = set(c for c, _ in combos)
    print(f"Found {len(companies)} clients with {len(combos)} Microsoft product subscriptions")
    print(f"Found {len(product_id_map)} unique products to price")

    # Authenticate with PAX8 and fetch live pricing
    client_id, client_secret = load_credentials()
    api_pricing = {}
    has_pricing = False

    if client_id and client_secret:
        session = authenticate(client_id, client_secret)
        if session:
            print(f"\nFetching live pricing from PAX8 API for {len(product_id_map)} products...")
            api_pricing = fetch_all_pricing(session, list(product_id_map.keys()))
            has_pricing = len(api_pricing) > 0
            if has_pricing:
                print(f"\nPricing available for {len(api_pricing)}/{len(product_id_map)} products")
            else:
                print("\nNo products had both monthly and annual pricing available")
    else:
        print("\nSkipping pricing — no PAX8 credentials configured")

    # Build timelines
    print("\nBuilding monthly timelines...")
    combo_info = build_timelines(summary_records, history_records)

    # Calculate metrics and recommendations
    print("Calculating recommendations...")
    results = {}
    unpriced = {}

    for key, info in combo_info.items():
        company, product = key
        timeline = info.get("timeline", {})
        current_count = info.get("current_count", 0)
        sku = info.get("sku", "")
        product_id = info.get("product_id", "")

        metrics = calculate_metrics(timeline, current_count)
        rec = generate_recommendations(metrics)

        # Look up pricing by product_id from API results
        price_info = api_pricing.get(product_id)

        if has_pricing and price_info is None:
            ukey = (product, sku, product_id)
            if ukey not in unpriced:
                reason = "No product ID" if not product_id else "No monthly+annual pricing from API"
                unpriced[ukey] = {"client_count": 0, "total_licenses": 0, "reason": reason}
            unpriced[ukey]["client_count"] += 1
            unpriced[ukey]["total_licenses"] += current_count

        savings = calculate_savings(rec, metrics, price_info)

        results[key] = {
            "sku": sku,
            "metrics": metrics,
            "rec": rec,
            "savings": savings,
        }

    if unpriced:
        print(f"Unpriced products: {len(unpriced)} — see 'Unpriced Products' tab")

    # Generate Excel
    print("\nGenerating Excel report...", end=" ", flush=True)
    out_wb = Workbook()
    write_recommendations_tab(out_wb, results, has_pricing)
    write_trends_tab(out_wb, results, combo_info)
    write_savings_summary_tab(out_wb, results, has_pricing)
    if unpriced:
        write_unpriced_tab(out_wb, unpriced)

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"license_optimization_{today}.xlsx"
    filepath = os.path.join(script_dir, filename)
    out_wb.save(filepath)
    print("Done.")

    abs_path = os.path.abspath(filepath)
    print(f"\nReport saved to: {abs_path}")

    if has_pricing:
        total_con = sum(r["savings"]["conservative"]["savings"] or 0 for r in results.values())
        total_mod = sum(r["savings"]["moderate"]["savings"] or 0 for r in results.values())
        total_agg = sum(r["savings"]["aggressive"]["savings"] or 0 for r in results.values())
        print(f"Portfolio savings potential: Conservative ${total_con:,.2f}/yr | "
              f"Moderate ${total_mod:,.2f}/yr | Aggressive ${total_agg:,.2f}/yr")


if __name__ == "__main__":
    main()
